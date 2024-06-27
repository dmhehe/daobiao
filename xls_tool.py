import re, os
import json
from openpyxl import load_workbook



ATTR_ROW = 6 #属性行，前6行都是属性行，没有数据的
variable_name_pattern = re.compile(r'^[a-zA-Z_]\w*$')

def getSheetName(xls_file_path):
    # 打开一个 xlsx 文件
    workbook = load_workbook(filename=xls_file_path, data_only=True)
    
    # 获取所有工作表的名字
    sheets_names = workbook.sheetnames
    
    # 定义一个正则表达式匹配符合变量命名规则的名称
    # 变量名可以使用字母、数字、下划线，但不能以数字开头
    variable_name_pattern = re.compile(r'^[a-zA-Z_][a-zA-Z0-9_]*$')
    
    # 过滤掉以 "Sheet" 开头和不符合变量命名规则的工作表名称
    filtered_sheet_names = [
        name for name in sheets_names
        if not name.startswith("Sheet") and variable_name_pattern.match(name)
    ]
    
    return filtered_sheet_names

#留下有用的部分
def filter_usedata(data):
    if len(data) < 5:
        raise Exception("表数据有问题1！！小于5行")

#    for i, list1 in enumerate(data):
#        print("4444444444444", i, list1)


    max_col = len(data[0])
    for i in range(max_col):
        list1 = []
        for j in range(5):
            list1.append(data[j][i].strip())
        if all(s == "" for s in list1):
            max_col = i
            break
        
    max_row = len(data)
    for i, row_data in enumerate(data):
        list1 = []
        for j in range(max_col):
            list1.append(row_data[j])
        if all(s == "" for s in list1):
            max_row = i
            break

    new_data = []

    for i in range(max_row):
        new_line = []
        new_data.append(new_line)
        for j in range(max_col):
            new_line.append(data[i][j])


    if len(new_data) < 5:
        raise Exception("表有效数据有问题2！！小于5行")

    return new_data



def read_sheet_data(xls_file_path, sheet_name):
    # 打开 xlsx 文件
    workbook = load_workbook(filename=xls_file_path, data_only=True)
    
    # 通过名称获取工作表
    sheet = workbook[sheet_name]
    
    # 读取工作表数据到二维字符串列表
    data = []
    for row in sheet.iter_rows(values_only=True):
        row_data = []
        for cell_value in row:
            # 处理数据类型，确保结果是字符串
            # print("ssssssssss", cell_value)
            if isinstance(cell_value, str):
                row_data.append(cell_value.strip())
            elif cell_value is None:
                row_data.append('')
            else:
                # 对于非字符串类型，如数字，转换为字符串
                row_data.append(str(cell_value).strip())
        data.append(row_data)
    
    data = filter_usedata(data)
    return data


#返回属性字典，‘1,a=2,b=3’
def getAttrDict(text, mainName="main"):
    text = text.replace("，", ",")
    data_dict = {}
    attr_pair_list = text.split(",")
    for attr_pair_text in attr_pair_list:
        txt1 = attr_pair_text.strip()
        if not txt1:
            continue
        
        if txt1.find("=") >= 0:
            list1 = txt1.split("=")
            data_dict[list1[0]] = list1[1]
        else:
            data_dict[mainName] = txt1

    return data_dict


def getNumList(text):
    text = text.replace("，", ",")
    return eval("["  + text + "]")


def getStringList(text):
    text = text.replace("，", ",")
    list1 = text.split(",")
    list2 = []
    for str1 in list1:
        list2.append(str1.strip())

    return list2


def convert_string_to_number(s):
    try:
        # 尝试将字符串转换为整数
        return int(s)
    except ValueError:
        try:
            # 如果转换为整数失败，尝试转换为浮点数
            return float(s)
        except ValueError:
            # 如果转换为整数和浮点数都失败，返回原始字符串
            return s






def to_lua(data, indent=0, line_data_list=[]):
    lua_str = ""
    indent_str = " " * (indent * 4)

    if isinstance(data, dict):
        if data in line_data_list:
            lua_str += "{ "
            for key, value in data.items():
                if isinstance(key, str):
                    lua_str += f'["{key}"] = {to_lua(value, 0, line_data_list)}, '
                else:
                    lua_str += f'[{key}] = {to_lua(value, 0, line_data_list)}, '
            lua_str += "}"
        else:
            lua_str += "{\n"
            for key, value in data.items():
                if isinstance(key, str):
                    lua_str += f'{indent_str}    ["{key}"] = {to_lua(value, indent + 1, line_data_list)},\n'
                else:
                    lua_str += f'{indent_str}    [{key}] = {to_lua(value, indent + 1, line_data_list)},\n'
            lua_str += indent_str + "}"
    elif isinstance(data, list):
        if data in line_data_list:
            lua_str += "{ "
            for item in data:
                lua_str += f'{to_lua(item, 0, line_data_list)}, '
            lua_str += "}"
        else:
            lua_str += "{\n"
            for item in data:
                lua_str += f'{indent_str}    {to_lua(item, indent + 1, line_data_list)},\n'
            lua_str += indent_str + "}"
    elif isinstance(data, str):
        lua_str += f'"{data}"'
    elif isinstance(data, (int, float)):
        lua_str += str(data)
    elif data is None:
        lua_str += "nil"
    else:
        raise TypeError(f"Unsupported data type: {type(data)}")

    return lua_str






g_Start_Flag = "//----&&----start----&&----"
g_End_Flag = "//----&&----end----&&----"
def writeFile(file_path, str1):
    content = ""
    # 检查文件是否存在
    if os.path.exists(file_path):
        # 打开并读取文件
        with open(file_path, 'r', encoding='utf-8') as file1:
            # 读取文件的全部内容
            content = file1.read()
    else:
        # 如果文件不存在，初始化默认内容
        content = f"{g_Start_Flag}\n{g_End_Flag}"

    # 查找标志位置
    start_idx = content.find(g_Start_Flag)
    end_idx = content.find(g_End_Flag)

    # 初始化文本部分
    start_text = ""
    end_text = ""

    # 确保标志存在并且索引有效
    if start_idx != -1:
        start_text = content[:start_idx]
    if end_idx != -1:
        end_text = content[end_idx + len(g_End_Flag):]

    # 构建最终文本
    text = start_text + g_Start_Flag + "\n" + str1 + "\n" + g_End_Flag + end_text


    directory = os.path.dirname(file_path)
    if not os.path.exists(directory):
        os.makedirs(directory)

    # 写入文件
    with open(file_path, 'w', encoding='utf-8') as file2:
        file2.write(text)

def writeJson(file_path, data):
    
    directory = os.path.dirname(file_path)
    if not os.path.exists(directory):
        os.makedirs(directory)
    
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)
        
        
def writeLua(file_path, data, startText = "", line_data_list=[]):
    str1 = startText + to_lua(data, 0, line_data_list)
    writeFile(file_path, str1)


